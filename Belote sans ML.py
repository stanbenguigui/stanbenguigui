import random
import openpyxl
import csv
import itertools
from collections import defaultdict
from collections import deque
import openpyxl


valeurs = ["7", "8", "9", "10", "Valet", "Dame", "Roi", "As"]
couleurs = ["Coeur", "Carreau", "Trèfle", "Pique"]
points_values = {"7": 0, "8": 0, "9": 0, "10": 10, "Valet": 2, "Dame": 3, "Roi": 4, "As": 11}
atout_points = {"7": 0, "8": 0, "9": 14, "10": 10, "Valet": 20, "Dame": 3, "Roi": 4, "As": 11}

class Carte:
    def __init__(self, valeur, couleur):
        self.valeur = valeur
        self.couleur = couleur

    def __repr__(self):
        return "{} de {}".format(self.valeur, self.couleur)

    def __eq__(self, other):
        if isinstance(other, Carte):
            return self.valeur == other.valeur and self.couleur == other.couleur
        return False

    def __hash__(self):
        return hash((self.valeur, self.couleur))

class Joueur:
    def __init__(self, nom, equipe):
        self.nom = nom
        self.cartes = []
        self.points = 0
        self.equipe = equipe
        self.indice_carte_actuelle = 0
        self.indice_joueur = 0  # Nouvel attribut pour l'indice du joueur


class Equipe:
    def __init__(self):
        self.points = 0
        self.belote = False
        self.dernier_pli = False

class Belote:
    def __init__(self):
        self.cartes = []
        for couleur in couleurs:
            for valeur in valeurs:
                self.cartes.append(Carte(valeur, couleur))
        self.joueurs = []

def init_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats Belote"
    ws.append(["Numéro de Partie", "Couleur Atout", "Carte Joueur 1", "Carte Joueur 2", "Carte Joueur 3", "Carte Joueur 4", "Points du Pli", "Gagnant du Pli", "Premier Joueur"])
    return wb, ws


def distribuer_cartes(joueurs, jeu_cartes, cartes_joueur1):
    # Convertir les cartes de jeu en un ensemble pour une manipulation facile
    set_jeu_cartes = set(jeu_cartes)

    # S'assurer que les cartes du joueur 1 sont retirées du pool de cartes
    cartes_restantes = set_jeu_cartes - set(cartes_joueur1)

    # Distribuer les cartes restantes aux autres joueurs
    for joueur in joueurs[1:]:
        joueur.cartes = random.sample(cartes_restantes, 8)
        cartes_restantes -= set(joueur.cartes)

    # Vérifier que les cartes du joueur 1 n'ont pas été redistribuées
    assert not any(carte in cartes_joueur1 for joueur in joueurs[1:] for carte in joueur.cartes), "Erreur: une carte du joueur 1 a été redistribuée."

def saisir_main_joueur1(jeu_cartes):
    abreviations = {
        "7": "7", "8": "8", "9": "9", "T": "10",
        "V": "Valet", "D": "Dame", "R": "Roi", "A": "As"
    }
    cartes_joueur1 = []
    print("Entrez les cartes de la main du joueur 1 (par exemple 7C pour 7 de Coeur) :")
    while len(cartes_joueur1) < 8:
        carte_saisie = input().upper()
        if len(carte_saisie) != 2 or carte_saisie[0] not in abreviations or carte_saisie[1] not in "CKTP":
            print("Format de carte invalide. Réessayez.")
            continue

        valeur = abreviations[carte_saisie[0]]
        couleur = {"C": "Coeur", "K": "Carreau", "T": "Trèfle", "P": "Pique"}[carte_saisie[1]]
        carte = Carte(valeur, couleur)

        if carte not in jeu_cartes:
            print("La carte saisie n'existe pas dans le jeu. Réessayez.")
        elif carte in cartes_joueur1:
            print("Cette carte a déjà été saisie. Réessayez.")
        else:
            cartes_joueur1.append(carte)
            if len(cartes_joueur1) == 8:
                break  # Sortie de la boucle une fois 8 cartes valides saisies
    
    return cartes_joueur1


class ModeleApprentissage:
    def __init__(self):
        # Un modèle d'apprentissage pour chaque couleur d'atout
        self.modeles = {
            "Coeur": defaultdict(int),
            "Carreau": defaultdict(int),
            "Trèfle": defaultdict(int),
            "Pique": defaultdict(int)
        }
        self.statistiques = defaultdict(int)

    def enregistrer_experience(self, etat, action, recompense, atout):
        etat_action = (str(etat), str(action))
        self.modeles[atout][etat_action] += recompense

    def predire_action(self, etat_jeu, atout):
        actions_possibles = etat_jeu['actions_possibles']
        meilleur_score = float('-inf')
        meilleure_action = None

        for action in actions_possibles:
            etat_action = (str(etat_jeu), str(action))
            score = self.modeles[atout].get(etat_action, 0)
            if score > meilleur_score:
                meilleur_score = score
                meilleure_action = action

        return meilleure_action

    def entrainer_modele(self, atout):
        # Calculer la moyenne des récompenses pour chaque action
        for etat_action, total_recompense in self.modeles[atout].items():
            nombre_occurrences = self.statistiques[etat_action]
            if nombre_occurrences > 0:
                self.modeles[atout][etat_action] = total_recompense / nombre_occurrences


def enregistrer_decision(joueur, carte, pli, points, atout, modele):
    etat = (tuple(joueur.cartes), tuple(pli), carte)
    modele.enregistrer_experience(etat, carte, points, atout)


def choisir_carte_ordinateur_ameliore(joueur, pli, atout, joueurs, modele):
    # Ajout de la vérification pour le premier pli de la partie
    if not pli and all(len(j.cartes) == 8 for j in joueurs):
        return random.choice(joueur.cartes)
    def est_maitre(carte, pli, atout):
        if not pli:
            return True
        if carte.couleur == atout:
            return all(c.couleur != atout or atout_points[carte.valeur] > atout_points[c.valeur] for c in pli)
        return carte.couleur == pli[0].couleur and all(c.couleur != atout and points_values[carte.valeur] > points_values[c.valeur] for c in pli)

    def a_coupe_adverse():
        for c in pli:
            if c.couleur == atout and joueurs[pli.index(c)].equipe != joueur.equipe:
                return True
        return False

    def coequipier_a_maitre():
        for i, carte in enumerate(pli):
            if joueurs[i].equipe == joueur.equipe and est_maitre(carte, pli, atout):
                return True
        return False

    # Vérifier si le joueur a encore des cartes
    if not joueur.cartes:
        return None

    cartes_jouables = [carte for carte in joueur.cartes if carte.couleur == (pli[0].couleur if pli else carte.couleur)]
    if not cartes_jouables:
        cartes_jouables = joueur.cartes

    # Sélection de la carte
    if a_coupe_adverse() or not coequipier_a_maitre():
        if not cartes_jouables:
            cartes_jouables = [carte for carte in joueur.cartes if carte.couleur != atout]  # Éviter de couper si possible
            if not cartes_jouables:
                cartes_jouables = joueur.cartes  # Si aucune autre option, utiliser toutes les cartes
        carte_choisie = min(cartes_jouables, key=lambda c: points_values[c.valeur])
    else:
        cartes_maitres = [c for c in cartes_jouables if est_maitre(c, pli, atout)]
        if cartes_maitres:
            carte_choisie = max(cartes_maitres, key=lambda c: points_values[c.valeur])
        else:
            carte_choisie = min(cartes_jouables, key=lambda c: points_values[c.valeur])

    # Retirez la carte choisie de la main du joueur
    joueur.cartes.remove(carte_choisie)
    return carte_choisie


def choisir_carte_pour_joueur1(joueur, pli, atout, joueurs):
    # Stratégie de base pour choisir la meilleure carte
    def est_maitre():
        for carte in reversed(pli):
            if carte.couleur == atout or carte.couleur == pli[0].couleur:
                return joueurs[pli.index(carte)].equipe == joueur.equipe
        return False

    def carte_la_plus_forte(cartes):
        return max(cartes, key=lambda c: points_values[c.valeur]) if cartes else None

    def carte_la_plus_faible(cartes):
        return min(cartes, key=lambda c: points_values[c.valeur]) if cartes else None

    # Vérifier si le joueur a encore des cartes
    if not joueur.cartes:
        return None

    cartes_jouables = [carte for carte in joueur.cartes if carte.couleur == pli[0].couleur] if pli else joueur.cartes

    if not cartes_jouables:
        atouts = [carte for carte in joueur.cartes if carte.couleur == atout]
        carte_choisie = carte_la_plus_faible(atouts) if not est_maitre() else carte_la_plus_forte(atouts)
    else:
        carte_choisie = carte_la_plus_forte(cartes_jouables) if est_maitre() else carte_la_plus_faible(cartes_jouables)

    if carte_choisie:
        joueur.cartes.remove(carte_choisie)  # Retirer la carte choisie de la main du joueur
        return carte_choisie

    return random.choice(joueur.cartes)  # Choix aléatoire si aucune autre option





def choisir_carte_ordinateur(joueur, pli, atout, joueurs):
    def est_maitre():
        for carte in reversed(pli):
            if carte.couleur == atout or carte.couleur == pli[0].couleur:
                return joueurs[pli.index(carte)].equipe == joueur.equipe
        return False

    def carte_la_plus_forte(cartes):
        return max(cartes, key=lambda c: points_values[c.valeur]) if cartes else None

    def carte_la_plus_faible(cartes):
        return min(cartes, key=lambda c: points_values[c.valeur]) if cartes else None

    # Vérifier si le joueur a encore des cartes
    if not joueur.cartes:
        return None

    cartes_jouables = [carte for carte in joueur.cartes if carte.couleur == pli[0].couleur] if pli else joueur.cartes

    if not cartes_jouables:
        atouts = [carte for carte in joueur.cartes if carte.couleur == atout]
        carte_choisie = carte_la_plus_faible(atouts) if not est_maitre() else carte_la_plus_forte(atouts)
    else:
        carte_choisie = carte_la_plus_forte(cartes_jouables) if est_maitre() else carte_la_plus_faible(cartes_jouables)

    if carte_choisie:
        joueur.cartes.remove(carte_choisie)  # Retirer la carte choisie de la main du joueur
        return carte_choisie

    return random.choice(joueur.cartes)  # Choix aléatoire si aucune autre option



def determiner_gagnant_pli(jeu, pli, atout):
    meilleur_carte = pli[0]
    gagnant = jeu.joueurs[0]

    for i in range(1, len(pli)):
        carte = pli[i]
        joueur = jeu.joueurs[i]

        if meilleur_carte.couleur == atout and carte.couleur == atout:
            if atout_points[carte.valeur] > atout_points[meilleur_carte.valeur]:
                meilleur_carte = carte
                gagnant = joueur
        elif carte.couleur == atout:
            meilleur_carte = carte
            gagnant = joueur
        elif carte.couleur == meilleur_carte.couleur and points_values[carte.valeur] > points_values[meilleur_carte.valeur]:
            meilleur_carte = carte
            gagnant = joueur

    return gagnant


def jouer_manche(joueur1_cartes, atout, equipe1, equipe2, ws, num_partie, modele):
    jeu = Belote()
    joueur1 = Joueur("Joueur 1", equipe1)
    joueur2 = Joueur("Ordinateur 1", equipe2)
    joueur3 = Joueur("Ordinateur 2", equipe1)
    joueur4 = Joueur("Ordinateur 3", equipe2)
    jeu.joueurs = [joueur1, joueur2, joueur3, joueur4]


    joueur1.cartes = joueur1_cartes.copy()
    distribuer_cartes(jeu.joueurs, jeu.cartes, joueur1_cartes)
    cartes_joueur1 = joueur1_cartes.copy()
    plis = []
    # Utilisez une copie des cartes pour éviter de modifier la liste originale
    cartes_joueur1 = joueur1_cartes.copy()

    for _ in range(8):
        pli = []
        cartes_du_pli = {}  # Dictionnaire pour stocker les cartes jouées par chaque joueur

        for joueur in jeu.joueurs:
            if joueur.nom == "Joueur 1":
                # Laisser l'utilisateur choisir une carte ou suggérer la meilleure carte
                carte = choisir_carte_ordinateur_ameliore(joueur, pli, atout, jeu.joueurs, modele)
            else:
                carte = choisir_carte_ordinateur_ameliore(joueur, pli, atout, jeu.joueurs, modele)
            pli.append(carte)
            cartes_du_pli[joueur.nom] = carte

        gagnant = determiner_gagnant_pli(jeu, pli, atout)

        # Réorganisation des joueurs
        while jeu.joueurs[0] != gagnant:
            jeu.joueurs.append(jeu.joueurs.pop(0))

        # Calcul des points
        points = sum(atout_points[carte.valeur] if carte.couleur == atout else points_values[carte.valeur] for carte in pli)
                
        # Réorganiser l'ordre des cartes selon l'ordre initial des joueurs
        cartes_pli_excel = [f"{cartes_du_pli[nom].valeur} de {cartes_du_pli[nom].couleur}" for nom in ["Joueur 1", "Ordinateur 1", "Ordinateur 2", "Ordinateur 3"]]
        ws.append([num_partie, atout] + cartes_pli_excel + [points, gagnant.nom, jeu.joueurs[0].nom])
    
        # Belote
        if not gagnant.equipe.belote:
            for i in range(len(pli) - 1):
                if pli[i].couleur == atout and pli[i].valeur == "Roi" and pli[i + 1].valeur == "Dame" or pli[i].valeur == "Dame" and pli[i + 1].valeur == "Roi":
                    gagnant.equipe.belote = True
                    gagnant.equipe.points += 20

        gagnant.equipe.points += points
        plis.append(pli)

    # Dernier pli
    gagnant_dernier_pli = determiner_gagnant_pli(jeu, plis[-1], atout)
    gagnant_dernier_pli.equipe.points += 10

    return gagnant.equipe.points, sum(joueur.points for joueur in jeu.joueurs)




def simuler_manche_pour_apprentissage(cartes_joueur1, atout, modele):
    # Créer une copie des cartes pour cette simulation particulière

    jeu = Belote()
    joueur1 = Joueur("Joueur 1", Equipe())
    joueur2 = Joueur("Ordinateur 1", Equipe())
    joueur3 = Joueur("Ordinateur 2", Equipe())
    joueur4 = Joueur("Ordinateur 3", Equipe())
    jeu.joueurs = [joueur1, joueur2, joueur3, joueur4]

    # Distribuer les cartes
    joueur1.cartes = cartes_joueur1
    # Distribuer les cartes en utilisant la copie
    distribuer_cartes(jeu.joueurs, jeu.cartes, cartes_joueur1)

    for _ in range(8):  # Chaque pli
        pli = []
        for joueur in jeu.joueurs:
            if joueur.nom == "Joueur 1":
                # Ici, on peut simuler une décision ou permettre une saisie manuelle
                carte_jouee = choisir_carte_ordinateur_ameliore(joueur, pli, atout, jeu.joueurs, modele)
            else:
                carte_jouee = choisir_carte_ordinateur_ameliore(joueur, pli, atout, jeu.joueurs, modele)
            pli.append(carte_jouee)
            enregistrer_decision(joueur, carte_jouee, pli, points_values[carte_jouee.valeur], atout, modele)

        # Déterminer le gagnant du pli et réorganiser les joueurs
        gagnant = determiner_gagnant_pli(jeu, pli, atout)
        while jeu.joueurs[0] != gagnant:
            jeu.joueurs.append(jeu.joueurs.pop(0))



def main():
    wb, ws = init_excel()
    modele = ModeleApprentissage()
    num_partie = 1
    jeu = Belote()

    if len(jeu.cartes) != 32:
        raise ValueError("Le jeu ne contient pas le bon nombre de cartes.")

    # Initialisation du modèle d'apprentissage
    modele = ModeleApprentissage()
    joueur1_cartes = saisir_main_joueur1(jeu.cartes)

    # Phase d'apprentissage
    for atout in ["Coeur", "Carreau", "Trèfle", "Pique"]:
        for _ in range(25000):  # Supposons que nous faisons 250 essais pour chaque couleur d'atout
            simuler_manche_pour_apprentissage(joueur1_cartes.copy(), atout, modele)
        modele.entrainer_modele(atout)  # Entraîner le modèle pour chaque couleur d'atout séparément

    # Phase de jeu principale
    num_partie = 1
    atout_couleurs = ["Coeur", "Carreau", "Trèfle", "Pique"]
    nb_essais = 1000
    resultats = defaultdict(lambda: defaultdict(int))
    lignes_ajoutees = 0
    for atout in atout_couleurs:
        if lignes_ajoutees >= 100000:
            break  # Arrêtez si la limite de ligne est atteinte
        for tranche in range(80, 170, 10):
            if lignes_ajoutees >= 100000:
                break
            nb_victoires = 0
            for _ in range(nb_essais):
                if lignes_ajoutees >= 100000:
                    break
                equipe1 = Equipe()
                equipe2 = Equipe()
                points_joueur1, _ = jouer_manche(joueur1_cartes, atout, equipe1, equipe2, ws, num_partie, modele)
                num_partie += 1
                lignes_ajoutees += 1
                if points_joueur1 >= tranche:
                    nb_victoires += 1
            pourcentage = (nb_victoires / nb_essais) * 100
            resultats[atout][tranche] = pourcentage

    # Affichage des statistiques de performance
#   print("\nStatistiques de performance du modèle :")
#   for stat, valeur in modele.statistiques.items():
#       print(f"{stat}: {valeur}")


    # Affichage et enregistrement des résultats
    print("\nRésultats pour chaque couleur d'atout (sur {} essais) :".format(nb_essais))
    print("{:<12}".format("Atout"), end="")
    for tranche in range(80, 170, 10):
        print("{:<10}".format(tranche), end="")
    print("")
    
    for atout in atout_couleurs:
        print("{:<12}".format(atout), end="")
        for tranche in range(80, 170, 10):
            print("{:<10.2f}%".format(resultats[atout][tranche]), end="")
        print("")

    # Enregistrement dans le fichier Excel
    nom_fichier = "R.B.Test33.xlsx"
    wb.save(nom_fichier)
    wb = openpyxl.load_workbook(nom_fichier)
    ws = wb.active
    print("\nAffichage des 50 premières lignes du fichier Excel:")
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
        print(row)

if __name__ == "__main__":
    main()